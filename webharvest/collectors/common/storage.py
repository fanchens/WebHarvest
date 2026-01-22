from __future__ import annotations

import csv
import json
from dataclasses import asdict, is_dataclass
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from .config import ProjectPaths
from .utils import safe_filename

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


class SaveFormat(str, Enum):
    """保存格式枚举"""
    JSON = "json"
    JSONL = "jsonl"
    CSV = "csv"
    EXCEL = "excel"
    TXT = "txt"


def _default_json(obj: Any):
    if is_dataclass(obj):
        return asdict(obj)
    return str(obj)


def _flatten_dict(d: Dict[str, Any], parent_key: str = "", sep: str = "_") -> Dict[str, Any]:
    """展平嵌套字典"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(_flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # 列表转为字符串
            items.append((new_key, json.dumps(v, ensure_ascii=False) if v else ""))
        else:
            items.append((new_key, v))
    return dict(items)


class JsonStorage:
    """
    JSON 保存（整文件：覆盖写）
    适合：一次性导出列表、配置快照等
    """

    def __init__(self, *, paths: Optional[ProjectPaths] = None):
        self.paths = (paths or ProjectPaths.detect()).ensure()

    def write(
        self,
        *,
        site: str,
        name: str,
        data: Any,
        subdir: str = "",
    ) -> Path:
        out_dir = self.paths.outputs_dir / (subdir or site)
        out_dir.mkdir(parents=True, exist_ok=True)
        fname = safe_filename(name)
        path = out_dir / f"{fname}.json"
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2, default=_default_json), encoding="utf-8")
        return path


class JsonlStorage:
    """
    JSONL 保存（按行追加）
    适合：采集过程持续写入、断点续写、流式落地
    """

    def __init__(self, *, paths: Optional[ProjectPaths] = None):
        self.paths = (paths or ProjectPaths.detect()).ensure()

    def append(
        self,
        *,
        site: str,
        name: str,
        item: Dict[str, Any],
        subdir: str = "",
    ) -> Path:
        out_dir = self.paths.outputs_dir / (subdir or site)
        out_dir.mkdir(parents=True, exist_ok=True)
        fname = safe_filename(name)
        path = out_dir / f"{fname}.jsonl"
        with path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(item, ensure_ascii=False, default=_default_json))
            f.write("\n")
        return path

    def append_many(
        self,
        *,
        site: str,
        name: str,
        items: Iterable[Dict[str, Any]],
        subdir: str = "",
    ) -> Path:
        out_dir = self.paths.outputs_dir / (subdir or site)
        out_dir.mkdir(parents=True, exist_ok=True)
        fname = safe_filename(name)
        path = out_dir / f"{fname}.jsonl"
        with path.open("a", encoding="utf-8") as f:
            for item in items:
                f.write(json.dumps(item, ensure_ascii=False, default=_default_json))
                f.write("\n")
        return path


class CsvStorage:
    """
    CSV 保存（表格格式，适合 Excel 打开）
    适合：结构化数据导出、数据分析
    """

    def __init__(self, *, paths: Optional[ProjectPaths] = None):
        self.paths = (paths or ProjectPaths.detect()).ensure()

    def write(
        self,
        *,
        site: str,
        name: str,
        data: List[Dict[str, Any]],
        subdir: str = "",
        flatten: bool = True,
    ) -> Path:
        """
        保存为 CSV
        - data: 字典列表
        - flatten: 是否展平嵌套字典（默认 True）
        """
        if not data:
            raise ValueError("数据为空，无法保存 CSV")

        out_dir = self.paths.outputs_dir / (subdir or site)
        out_dir.mkdir(parents=True, exist_ok=True)
        fname = safe_filename(name)
        path = out_dir / f"{fname}.csv"

        # 展平数据（如果需要）
        rows = [_flatten_dict(item) if flatten else item for item in data]

        # 收集所有字段名（按第一个字典的键顺序）
        fieldnames = list(rows[0].keys()) if rows else []

        with path.open("w", newline="", encoding="utf-8-sig") as f:  # utf-8-sig 让 Excel 正确识别中文
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        return path


class ExcelStorage:
    """
    Excel 保存（.xlsx 格式）
    适合：复杂表格、多工作表、格式化需求
    """

    def __init__(self, *, paths: Optional[ProjectPaths] = None):
        if not HAS_EXCEL:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")
        self.paths = (paths or ProjectPaths.detect()).ensure()

    def write(
        self,
        *,
        site: str,
        name: str,
        data: List[Dict[str, Any]],
        subdir: str = "",
        flatten: bool = True,
        sheet_name: str = "Sheet1",
    ) -> Path:
        """
        保存为 Excel
        - data: 字典列表
        - flatten: 是否展平嵌套字典（默认 True）
        - sheet_name: 工作表名称
        """
        if not data:
            raise ValueError("数据为空，无法保存 Excel")

        out_dir = self.paths.outputs_dir / (subdir or site)
        out_dir.mkdir(parents=True, exist_ok=True)
        fname = safe_filename(name)
        path = out_dir / f"{fname}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 展平数据
        rows = [_flatten_dict(item) if flatten else item for item in data]
        fieldnames = list(rows[0].keys()) if rows else []

        # 写入表头（加粗、居中）
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.alignment = header_alignment

        # 写入数据
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, field in enumerate(fieldnames, start=1):
                value = row_data.get(field, "")
                # 处理复杂类型
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自动调整列宽
        for col_idx, field in enumerate(fieldnames, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(field)),
                max(len(str(row_data.get(field, ""))) for row_data in rows),
                10,
            )
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        wb.save(path)
        return path


class TxtStorage:
    """
    TXT 保存（纯文本，自定义分隔符）
    适合：简单列表、日志式输出、自定义格式
    """

    def __init__(self, *, paths: Optional[ProjectPaths] = None):
        self.paths = (paths or ProjectPaths.detect()).ensure()

    def write(
        self,
        *,
        site: str,
        name: str,
        data: List[Dict[str, Any]],
        subdir: str = "",
        separator: str = "\t",  # 默认 Tab 分隔（适合 Excel 打开）
        flatten: bool = True,
    ) -> Path:
        """
        保存为 TXT
        - data: 字典列表
        - separator: 字段分隔符（默认 Tab，可用逗号、竖线等）
        - flatten: 是否展平嵌套字典（默认 True）
        """
        if not data:
            raise ValueError("数据为空，无法保存 TXT")

        out_dir = self.paths.outputs_dir / (subdir or site)
        out_dir.mkdir(parents=True, exist_ok=True)
        fname = safe_filename(name)
        path = out_dir / f"{fname}.txt"

        rows = [_flatten_dict(item) if flatten else item for item in data]
        fieldnames = list(rows[0].keys()) if rows else []

        with path.open("w", encoding="utf-8") as f:
            # 写入表头
            f.write(separator.join(fieldnames) + "\n")
            # 写入数据
            for row_data in rows:
                values = [str(row_data.get(field, "")) for field in fieldnames]
                f.write(separator.join(values) + "\n")

        return path


class DataExporter:
    """
    统一数据导出器（根据格式自动选择存储方式）
    用法：
        exporter = DataExporter()
        exporter.save(site="xiaohongshu", name="notes", data=[...], format=SaveFormat.EXCEL)
    """

    def __init__(self, *, paths: Optional[ProjectPaths] = None):
        self.paths = paths or ProjectPaths.detect().ensure()
        self._json_storage = JsonStorage(paths=self.paths)
        self._jsonl_storage = JsonlStorage(paths=self.paths)
        self._csv_storage = CsvStorage(paths=self.paths)
        self._txt_storage = TxtStorage(paths=self.paths)
        self._excel_storage = ExcelStorage(paths=self.paths) if HAS_EXCEL else None

    def save(
        self,
        *,
        site: str,
        name: str,
        data: List[Dict[str, Any]],
        format: SaveFormat = SaveFormat.JSON,
        subdir: str = "",
        **kwargs,
    ) -> Path:
        """
        统一保存接口
        - site: 站点标识（如 "xiaohongshu"）
        - name: 文件名（不含扩展名）
        - data: 数据列表（字典列表）
        - format: 保存格式（SaveFormat 枚举）
        - subdir: 子目录（可选）
        - **kwargs: 传递给具体存储类的额外参数
        """
        if format == SaveFormat.JSON:
            return self._json_storage.write(site=site, name=name, data=data, subdir=subdir)
        elif format == SaveFormat.JSONL:
            # JSONL 需要逐条追加
            if data:
                first = data[0]
                self._jsonl_storage.append(site=site, name=name, item=first, subdir=subdir)
                if len(data) > 1:
                    self._jsonl_storage.append_many(site=site, name=name, items=data[1:], subdir=subdir)
            return self.paths.outputs_dir / (subdir or site) / f"{safe_filename(name)}.jsonl"
        elif format == SaveFormat.CSV:
            return self._csv_storage.write(site=site, name=name, data=data, subdir=subdir, **kwargs)
        elif format == SaveFormat.EXCEL:
            if not self._excel_storage:
                raise ImportError("Excel 格式需要安装 openpyxl: pip install openpyxl")
            return self._excel_storage.write(site=site, name=name, data=data, subdir=subdir, **kwargs)
        elif format == SaveFormat.TXT:
            return self._txt_storage.write(site=site, name=name, data=data, subdir=subdir, **kwargs)
        else:
            raise ValueError(f"不支持的格式: {format}")

    @staticmethod
    def get_supported_formats() -> List[str]:
        """获取支持的格式列表"""
        formats = [SaveFormat.JSON.value, SaveFormat.JSONL.value, SaveFormat.CSV.value, SaveFormat.TXT.value]
        if HAS_EXCEL:
            formats.append(SaveFormat.EXCEL.value)
        return formats


